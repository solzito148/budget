from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "CIUDAD/Gastos_CIUDAD_1132_2026.xlsx"
OUTPUT = ROOT / "SOL/Gastos_SOL_2026.xlsx"
MONTHS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
MONTH_NUM = {name.upper(): number for number, name in MONTHS.items()}

NAVY = "17365D"
WHITE = "FFFFFF"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
THIN = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_usd(comment):
    match = re.search(r"US\$\s*([\d.,]+)", str(comment or ""), re.I)
    if not match:
        return None
    raw = match.group(1)
    try:
        if "," in raw and "." in raw:
            return float(raw.replace(".", "").replace(",", "."))
        if "," in raw:
            return float(raw.replace(",", "."))
        return float(raw)
    except ValueError:
        return None


def normalized_category(concept, comment, medium, kind):
    text = " ".join(str(value or "") for value in (concept, comment, medium)).upper()
    if kind == "ALMACÉN SOL":
        if (
            "PEDIDOS YA · SUPERMERCADO" in text
            or "SUPERMERCADO" in text
            or any(value in text for value in ("COTO", "CARREFOUR", "DÍA", "DIA ONLINE", "SUPERDIA"))
        ):
            return "SUPERMERCADO"
        if (
            "RAMON · VETERINARIO" in text
            or "RAMÓN · VETERINARIO" in text
            or "RAMON · VETERINARIA" in text
            or "RAMÓN · VETERINARIA" in text
            or "PERRO VERDE" in text
            or ("VALERIA" in text and "MELE" in text)
            or "LEOCAN" in text
        ):
            return "RAMON · VETERINARIO"
        if any(
            value in text
            for value in (
                "RAMON · COMIDA",
                "RAMÓN · COMIDA",
                "COMIDA · RAMON",
                "COMIDA · RAMÓN",
                "DR BARF",
                "KALBB",
            )
        ):
            return "RAMON · COMIDA"
        if "RAMON · OTROS" in text or "RAMÓN · OTROS" in text:
            return "RAMON · OTROS"
        if "HOTEL MADERO" in text or "ENTRETENIMIENTO · SALIDAS COMIDA" in text or "SALIDAS COMIDA" in text:
            return "ENTRETENIMIENTO · SALIDAS COMIDA"
        if "ROLDAN" in text or "ENTRETENIMIENTO · SOCIAL" in text:
            return "ENTRETENIMIENTO SOCIAL"
        if "COSI MI PIACE" in text or "ENTRETENIMIENTO · CENAS" in text:
            return "ENTRETENIMIENTO CENAS"
        if "MOVISTAR ARENA" in text or "ENTRETENIMIENTO · RECITALES" in text or "RECITALES" in text:
            return "ENTRETENIMIENTO · RECITALES"
        if "PLATEANET" in text or "ENTRETENIMIENTO · TEATRO" in text:
            return "ENTRETENIMIENTO · TEATRO"
        if "HOYTS" in text or "ENTRETENIMIENTO · CINE" in text:
            return "ENTRETENIMIENTO · CINE"
        if "ENERFIT" in text or "BARRITAS PROTEICAS" in text:
            return "COMIDA · BARRITAS PROTEICAS"
        if "COMBUSTIBLE" in text or any(value in text for value in ("SHELL", "YPF", "GENERACION 20")):
            return "COMBUSTIBLE"
        if "STREAMING TV" in text or "DIRECTV" in text:
            return "STREAMING TV"
        if "ENTRETENIMIENTO" in text:
            return "ENTRETENIMIENTO"
        if "FARMACIA" in text or "FARMACITY" in text:
            return "FARMACIA"
        if "REPUESTOS" in text or "SODASTREAM" in text:
            return "REPUESTOS"
        return "COMIDA"

    if any(value in text for value in ("PERCEPCIÓN USD", "PERCEPCION USD", "RG 5617", "RG 4240")):
        return "PERCEPCIONES USD"
    if "AVA FIDEICOMISO" in text or "FIDEICOMISO AVA" in text:
        return "AVA FIDEICOMISO"
    if any(value in text for value in ("EVENTOS DEPORTIVOS - RIVER", "CUOTA SOCIAL CARP", "MONUMENTAL")):
        return "EVENTOS DEPORTIVOS - RIVER"
    if (
        "TRAÍDO DE MIAMI" in text
        or "TRAIDO DE MIAMI" in text
        or "PAGOS VIRTUALES" in text
        or "SERVICIOS DE PAGOS VIRTUALES" in text
    ):
        return "TRAÍDO DE MIAMI"
    if (
        "USA BOX" in text
        or "CARENZO" in text
        or "FEDERICO MARTINEZ ESCARIZ" in text
        or "ESCARIZ" in text
    ):
        return "USA BOX"
    if "FARMACIA SALUD" in text or "FARMACIA MAURE" in text or "FARMACIA SELMA" in text:
        return "FARMACIA SALUD"
    if "PROMARINE" in text or "SALUD · VITAMINAS" in text or "SALUD · VITAMIN" in text:
        return "SALUD · Vitaminas"
    if (
        "SALUD" in text
        or any(value in text for value in ("LABORATORIO", "BIODEPOT", "ENERFIT", "MADMUSCLES", "BOSSRECOVERY"))
    ):
        return "SALUD"
    if (
        "ROPA" in text
        or any(
            value in text
            for value in (
                "CEFERINA", "PANDORA", "ZARA", "GUCCI", "VUITTON",
                "ADIDAS", "AWADA", "ZADIG", "BIMBAYLOLA", "TIENDARIVER",
                "LEVIS", "JUSTA OSADIA",
            )
        )
    ):
        return "ROPA"
    if any(value in text for value in ("CABIFY", "UBER", "DIDI", "EBANX", "EMOVA")):
        return "TRANSPORTE"
    if "AUTO" in text or re.search(r"\bACA\b", text):
        return "AUTO"
    if any(value in text for value in ("MASCOTA", "FELIX", "FIDO")):
        return "MASCOTAS"
    if "MERCADO LIBRE" in text or "MERCADOLIBRE" in text or re.search(r"\bML\b", text):
        return "MERCADO LIBRE"
    if (
        "SUSCRIPCIONES · IA" in text
        or any(
            value in text
            for value in (
                "OPENAI", "CHATGPT", "GPT",
                "GOOGLE ONE", "GEMINI",
                "VERCEL", "CURSOR", "CLAUDE", "ANTHROPIC",
            )
        )
    ):
        return "SUSCRIPCIONES · IA"
    if "SUSCRIPCIONES · APPLE" in text or "APPLE" in text:
        return "SUSCRIPCIONES · APPLE"
    if any(value in text for value in ("WORDPRESS", "WATCHFACE")):
        return "SUSCRIPCIONES Y TECNOLOGÍA"
    if any(value in text for value in ("TUENTI", "MOVISTAR", "TELECOM")):
        return "TELECOMUNICACIONES"
    if any(value in text for value in ("EDUCACIÓN", "EDUCACION", "UNIVE FASTA", "UADE", "UNIVERSIDAD")):
        return "EDUCACIÓN"
    if "DONACI" in text:
        return "DONACIONES"
    if "ZURICH" in text:
        return "SEGUROS / RETIRO"
    if "AMAZON" in text:
        return "COMPRAS ONLINE"
    if "COMIDA · OFICINA" in text or re.search(r"\bFEED\b", text):
        return "COMIDA · OFICINA"
    if "ENTRETENIMIENTO · COMIDAS" in text or "OPORTO" in text:
        return "ENTRETENIMIENTO · COMIDAS"
    if "PELUQUER" in text or "NEWPLAMIER" in text:
        return "PELUQUERÍA"
    if (
        "VESTIMENTA" in text
        or "LONGCHAMP" in text
        or "ITSCLASSIC" in text
        or ("ANALIA" in text and "ROAY" in text)
        or ("JAZMIN" in text and "CHEBAR" in text)
        or ("AYLEN" in text and "DIFEO" in text)
    ):
        return "VESTIMENTA"
    if "FUNERAL" in text or ("JARDIN" in text and "PILAR" in text):
        return "FUNERAL"

    if "TELECOMUNICACIONES" in text or "TUENTI" in text:
        return "TELECOMUNICACIONES"
    if "SEGUROS" in text or "ZURICH" in text:
        return "SEGUROS · RETIRO"
    if "DONACION" in text or "DONACIÓN" in text or "SOPLO DE VIDA" in text:
        return "DONACIONES"
    if "TRANSFERENCIAS" in text:
        return "TRANSFERENCIAS"
    if "COMPRAS" in text:
        return "COMPRAS"
    if text.startswith("HOGAR") or "TEXTILHOGAR" in text or "SAPHIRUS" in text:
        return "HOGAR"
    if "TECNOLOG" in text or "TODOVISION" in text or "MOPHIE" in text or "SMARTCELL" in text:
        return "TECNOLOGÍA"
    if "ENVÍOS" in text or "ENVIOS" in text or "CORREO ARGENTINO" in text:
        return "ENVÍOS"
    if "LIBRER" in text or "PAPELERA" in text:
        return "LIBRERÍA"
    if "REGALOS" in text:
        return "REGALOS"

    return "OTROS PERSONALES"


def source_name(medium, comment):
    text = " ".join(str(value or "") for value in (medium, comment)).upper()
    if "TC MERCADO PAGO" in text:
        return "TC Mercado Pago"
    if "BBVA" in text:
        return "BBVA"
    if "MERCADO PAGO" in text:
        return "Mercado Pago"
    if "AVA" in text:
        return "AVA"
    return str(medium or "Sin identificar")


def year_month(item):
    if item["fecha"]:
        return item["fecha"].strftime("%Y-%m")
    source_month = str(item.get("mes_fuente") or "")
    if source_month.upper() in MONTH_NUM:
        return f"2026-{MONTH_NUM[source_month.upper()]:02d}"
    match = re.match(r"([A-Za-zÁÉÍÓÚáéíóú]+)\s+(\d{4})", source_month)
    if match and match.group(1).upper() in MONTH_NUM:
        return f"{match.group(2)}-{MONTH_NUM[match.group(1).upper()]:02d}"
    return "SIN FECHA"


def write_header(sheet, row, headers):
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row, column, header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def add_table(sheet, start_row, end_row, end_column, name):
    if end_row <= start_row:
        return
    table = Table(
        displayName=name,
        ref=f"A{start_row}:{get_column_letter(end_column)}{end_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def title(sheet, text, subtitle):
    sheet["A1"] = text
    sheet["A1"].font = Font(size=18, bold=True, color=WHITE)
    sheet["A1"].fill = HEADER_FILL
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(italic=True, color="666666")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)


def collect_sources():
    workbook = load_workbook(SOURCE, data_only=False)
    personal_sheet = workbook["Gasto Personal Sol"]
    warehouse_sheet = workbook["Almacen Detalle"]
    cashflow_sheet = workbook["Cashflow Cuotas BBVA"]
    compromisos_sheet = workbook["Cashflow Compromisos"] if "Cashflow Compromisos" in workbook.sheetnames else None

    personal = []
    projected = []
    for row in range(22, personal_sheet.max_row + 1):
        concept = personal_sheet.cell(row, 3).value
        amount = personal_sheet.cell(row, 4).value
        medium = personal_sheet.cell(row, 5).value
        comment = personal_sheet.cell(row, 6).value
        if concept is None or amount is None:
            continue
        # Skip summary labels in columns A/B that leak into concept area
        if str(concept).strip().upper() in {"CONCEPTO", "MONTO", "MEDIO", "COMENTARIO"}:
            continue
        try:
            ars = float(amount) if amount is not None else None
        except (TypeError, ValueError):
            continue
        movement_date = to_date(personal_sheet.cell(row, 1).value)
        is_projected = "PROYECTADO" in str(comment or "").upper()
        source_month = personal_sheet.cell(row, 2).value
        if isinstance(source_month, str) and source_month.startswith("="):
            source_month = None
        item = {
            "fecha": movement_date,
            "mes_fuente": source_month,
            "tipo": "GASTO PERSONAL",
            "categoria": normalized_category(concept, comment, medium, "GASTO PERSONAL"),
            "concepto": str(concept),
            "ars": ars,
            "usd": parse_usd(comment),
            "estado": "PROYECTADO" if is_projected else ("REVISAR FECHA" if movement_date is None else "REAL"),
            "medio": medium,
            "fuente": source_name(medium, comment),
            "comentario": comment,
            "origen": "CIUDAD / Gasto Personal Sol",
            "fila_origen": row,
        }
        item["anio_mes"] = year_month(item)
        (projected if is_projected else personal).append(item)

    warehouse = []
    for row in range(4, warehouse_sheet.max_row + 1):
        concept = warehouse_sheet.cell(row, 3).value
        amount = warehouse_sheet.cell(row, 4).value
        if concept is None or amount is None:
            continue
        try:
            ars = float(amount)
        except (TypeError, ValueError):
            continue
        movement_date = to_date(warehouse_sheet.cell(row, 2).value)
        comment = warehouse_sheet.cell(row, 7).value
        item = {
            "fecha": movement_date,
            "mes_fuente": warehouse_sheet.cell(row, 1).value,
            "tipo": "ALMACÉN SOL",
            "categoria": normalized_category(concept, comment, None, "ALMACÉN SOL"),
            "concepto": str(concept),
            "ars": ars,
            "usd": parse_usd(comment),
            "estado": "REVISAR FECHA" if movement_date is None else "REAL",
            "medio": None,
            "fuente": source_name(None, comment),
            "comentario": comment,
            "origen": "CIUDAD / Almacen Detalle",
            "fila_origen": row,
        }
        item["anio_mes"] = year_month(item)
        warehouse.append(item)

    cashflow = []
    for row in range(2, cashflow_sheet.max_row + 1):
        status = cashflow_sheet.cell(row, 10).value
        if status not in ("PAGADO", "PENDIENTE"):
            continue
        cashflow.append({
            "fuente": "BBVA TC",
            "compra": to_date(cashflow_sheet.cell(row, 1).value),
            "comercio": cashflow_sheet.cell(row, 2).value,
            "categoria": cashflow_sheet.cell(row, 3).value,
            "tarjeta": cashflow_sheet.cell(row, 4).value,
            "cuota": cashflow_sheet.cell(row, 5).value,
            "de": cashflow_sheet.cell(row, 6).value,
            "monto": float(cashflow_sheet.cell(row, 7).value or 0),
            "mes_cobro": cashflow_sheet.cell(row, 8).value,
            "anio_mes": cashflow_sheet.cell(row, 9).value,
            "estado": status,
            "comentario": cashflow_sheet.cell(row, 11).value,
            "responsable": "Sol",
        })

    no_gasto = []
    if "MP No Gasto" in workbook.sheetnames:
        no_gasto_sheet = workbook["MP No Gasto"]
        for row in range(24, no_gasto_sheet.max_row + 1):
            tipo = no_gasto_sheet.cell(row, 3).value
            amount = no_gasto_sheet.cell(row, 5).value
            if tipo is None or amount is None:
                continue
            try:
                monto = float(amount)
            except (TypeError, ValueError):
                continue
            no_gasto.append({
                "fecha": to_date(no_gasto_sheet.cell(row, 1).value),
                "mes_fuente": no_gasto_sheet.cell(row, 2).value,
                "tipo": str(tipo),
                "concepto": no_gasto_sheet.cell(row, 4).value,
                "monto": monto,
                "moneda": no_gasto_sheet.cell(row, 6).value,
                "medio": no_gasto_sheet.cell(row, 7).value,
                "comentario": no_gasto_sheet.cell(row, 8).value,
                "fila_origen": row,
            })

    compromisos = []
    if compromisos_sheet is not None:
        for row in range(2, compromisos_sheet.max_row + 1):
            fuente = compromisos_sheet.cell(row, 1).value
            status = compromisos_sheet.cell(row, 10).value
            if fuente not in ("BBVA TC", "AVA Fideicomiso") or status not in ("PAGADO", "PENDIENTE"):
                continue
            monto_raw = compromisos_sheet.cell(row, 7).value
            try:
                monto = float(monto_raw) if monto_raw is not None else None
            except (TypeError, ValueError):
                monto = None
            compromisos.append({
                "fuente": fuente,
                "compra": to_date(compromisos_sheet.cell(row, 2).value),
                "comercio": compromisos_sheet.cell(row, 3).value,
                "categoria": compromisos_sheet.cell(row, 4).value,
                "tarjeta": None,
                "cuota": compromisos_sheet.cell(row, 5).value,
                "de": compromisos_sheet.cell(row, 6).value,
                "monto": monto,
                "mes_cobro": compromisos_sheet.cell(row, 8).value,
                "anio_mes": compromisos_sheet.cell(row, 9).value,
                "estado": status,
                "comentario": compromisos_sheet.cell(row, 11).value,
                "responsable": compromisos_sheet.cell(row, 12).value,
            })
    return personal, warehouse, projected, cashflow, compromisos, no_gasto


def build():
    personal, warehouse, projected, cashflow, compromisos, no_gasto = collect_sources()
    actual = personal + warehouse
    actual_2026 = [item for item in actual if item["anio_mes"].startswith("2026-")]
    personal_2026 = [item for item in personal if item["anio_mes"].startswith("2026-")]
    warehouse_2026 = [item for item in warehouse if item["anio_mes"].startswith("2026-")]

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Resumen")
    movements = workbook.create_sheet("Movimientos Sol")
    personal_ws = workbook.create_sheet("Gastos Personales")
    warehouse_ws = workbook.create_sheet("Almacén Sol")
    cashflow_ws = workbook.create_sheet("Cashflow Cuotas BBVA")
    compromisos_ws = workbook.create_sheet("Cashflow Compromisos")
    no_gasto_ws = workbook.create_sheet("MP No Gasto")
    categories_ws = workbook.create_sheet("Categorías")
    perceptions_ws = workbook.create_sheet("Percepciones USD")
    control_ws = workbook.create_sheet("Control y fuentes")

    movement_headers = [
        "ID", "Fecha", "Año-Mes", "Mes fuente", "Tipo", "Categoría",
        "Concepto", "Monto ARS", "Monto USD", "Estado", "Medio", "Fuente",
        "Comentario", "Origen", "Fila origen",
    ]
    write_header(movements, 1, movement_headers)
    ordered_actual = sorted(
        actual,
        key=lambda item: (item["fecha"] or date(1900, 1, 1), item["tipo"], item["concepto"]),
    )
    for index, item in enumerate(ordered_actual, 1):
        row = index + 1
        values = [
            index, item["fecha"], item["anio_mes"], item["mes_fuente"],
            item["tipo"], item["categoria"], item["concepto"], item["ars"],
            item["usd"], item["estado"], item["medio"], item["fuente"],
            item["comentario"], item["origen"], item["fila_origen"],
        ]
        for column, value in enumerate(values, 1):
            movements.cell(row, column, value).border = THIN
        if item["fecha"]:
            movements.cell(row, 2).number_format = "YYYY-MM-DD"
        movements.cell(row, 8).number_format = "#,##0"
        movements.cell(row, 9).number_format = "#,##0.00"
        if item["estado"] == "REVISAR FECHA":
            movements.cell(row, 10).fill = PatternFill("solid", fgColor=YELLOW)
    add_table(movements, 1, len(actual) + 1, len(movement_headers), "TablaMovimientosSol")

    personal_headers = [
        "Fecha", "Año-Mes", "Mes fuente", "Categoría", "Concepto",
        "Monto ARS", "Monto USD", "Estado", "Medio", "Fuente",
        "Comentario", "Fila origen",
    ]
    write_header(personal_ws, 1, personal_headers)
    all_personal = sorted(
        personal + projected,
        key=lambda item: (item["fecha"] or date(1900, 1, 1), item["estado"], item["concepto"]),
    )
    for row, item in enumerate(all_personal, 2):
        values = [
            item["fecha"], item["anio_mes"], item["mes_fuente"], item["categoria"],
            item["concepto"], item["ars"], item["usd"], item["estado"],
            item["medio"], item["fuente"], item["comentario"], item["fila_origen"],
        ]
        for column, value in enumerate(values, 1):
            personal_ws.cell(row, column, value).border = THIN
        if item["fecha"]:
            personal_ws.cell(row, 1).number_format = "YYYY-MM-DD"
        personal_ws.cell(row, 6).number_format = "#,##0"
        personal_ws.cell(row, 7).number_format = "#,##0.00"
        if item["estado"] == "PROYECTADO":
            personal_ws.cell(row, 8).fill = PatternFill("solid", fgColor=YELLOW)
    add_table(personal_ws, 1, len(all_personal) + 1, len(personal_headers), "TablaGastosPersonales")

    warehouse_headers = [
        "Fecha", "Año-Mes", "Mes fuente", "Categoría", "Detalle",
        "Importe Sol ARS", "Monto USD", "Estado", "Fuente", "Comentario", "Fila origen",
    ]
    write_header(warehouse_ws, 1, warehouse_headers)
    for row, item in enumerate(
        sorted(warehouse, key=lambda value: (value["fecha"] or date(1900, 1, 1), value["concepto"])),
        2,
    ):
        values = [
            item["fecha"], item["anio_mes"], item["mes_fuente"], item["categoria"],
            item["concepto"], item["ars"], item["usd"], item["estado"],
            item["fuente"], item["comentario"], item["fila_origen"],
        ]
        for column, value in enumerate(values, 1):
            warehouse_ws.cell(row, column, value).border = THIN
        if item["fecha"]:
            warehouse_ws.cell(row, 1).number_format = "YYYY-MM-DD"
        warehouse_ws.cell(row, 6).number_format = "#,##0"
        warehouse_ws.cell(row, 7).number_format = "#,##0.00"
    add_table(warehouse_ws, 1, len(warehouse) + 1, len(warehouse_headers), "TablaAlmacenSol")

    cashflow_headers = [
        "Compra", "Comercio", "Categoría", "Tarjeta", "Cuota", "De",
        "Monto cuota ARS", "Mes cobro", "Año-Mes", "Estado", "Comentario",
    ]
    write_header(cashflow_ws, 1, cashflow_headers)
    for row, item in enumerate(cashflow, 2):
        values = [
            item["compra"], item["comercio"], item["categoria"], item["tarjeta"],
            item["cuota"], item["de"], item["monto"], item["mes_cobro"],
            item["anio_mes"], item["estado"], item["comentario"],
        ]
        for column, value in enumerate(values, 1):
            cashflow_ws.cell(row, column, value).border = THIN
        if item["compra"]:
            cashflow_ws.cell(row, 1).number_format = "YYYY-MM-DD"
        cashflow_ws.cell(row, 7).number_format = "#,##0"
        cashflow_ws.cell(row, 10).fill = PatternFill(
            "solid", fgColor=YELLOW if item["estado"] == "PENDIENTE" else GREEN
        )
    add_table(cashflow_ws, 1, len(cashflow) + 1, len(cashflow_headers), "TablaCashflowBBVA")

    compromisos_headers = [
        "Fuente", "Compra / Vencimiento", "Concepto", "Categoría", "Cuota", "De",
        "Monto ARS", "Mes", "Año-Mes", "Estado", "Comentario", "Responsable",
    ]
    write_header(compromisos_ws, 1, compromisos_headers)
    for row, item in enumerate(compromisos, 2):
        values = [
            item["fuente"], item["compra"], item["comercio"], item["categoria"],
            item["cuota"], item["de"], item["monto"], item["mes_cobro"],
            item["anio_mes"], item["estado"], item["comentario"], item["responsable"],
        ]
        for column, value in enumerate(values, 1):
            compromisos_ws.cell(row, column, value).border = THIN
        if item["compra"]:
            compromisos_ws.cell(row, 2).number_format = "YYYY-MM-DD"
        if item["monto"] is not None:
            compromisos_ws.cell(row, 7).number_format = "#,##0"
        compromisos_ws.cell(row, 10).fill = PatternFill(
            "solid", fgColor=YELLOW if item["estado"] == "PENDIENTE" else GREEN
        )
    if compromisos:
        add_table(compromisos_ws, 1, len(compromisos) + 1, len(compromisos_headers), "TablaCashflowCompromisos")

    no_gasto_headers = [
        "Fecha", "Mes", "Tipo", "Concepto", "Monto", "Moneda", "Medio", "Comentario", "Fila origen",
    ]
    write_header(no_gasto_ws, 1, no_gasto_headers)
    ordered_no_gasto = sorted(no_gasto, key=lambda item: (item["fecha"] or date(1900, 1, 1), item["tipo"]))
    for row, item in enumerate(ordered_no_gasto, 2):
        values = [
            item["fecha"], item["mes_fuente"], item["tipo"], item["concepto"],
            item["monto"], item["moneda"], item["medio"], item["comentario"],
            item["fila_origen"],
        ]
        for column, value in enumerate(values, 1):
            no_gasto_ws.cell(row, column, value).border = THIN
        if item["fecha"]:
            no_gasto_ws.cell(row, 1).number_format = "YYYY-MM-DD"
        no_gasto_ws.cell(row, 5).number_format = "#,##0"
    if no_gasto:
        add_table(no_gasto_ws, 1, len(no_gasto) + 1, len(no_gasto_headers), "TablaMPNoGasto")
    no_gasto_ws.column_dimensions["A"].width = 12
    no_gasto_ws.column_dimensions["C"].width = 20
    no_gasto_ws.column_dimensions["D"].width = 26
    no_gasto_ws.column_dimensions["H"].width = 60

    summary.sheet_view.showGridLines = False
    title(
        summary,
        "Gastos Sol — Consolidado 2026",
        "Archivo maestro: gastos personales, Almacén Sol, percepciones y cashflow (BBVA + AVA Fideicomiso)",
    )
    write_header(summary, 4, ["Indicador", "Monto ARS", "Detalle"])
    ava_pagado_sol = sum(
        item["monto"] or 0 for item in compromisos
        if item["fuente"] == "AVA Fideicomiso" and item["estado"] == "PAGADO"
    )
    ava_pendiente_n = sum(
        1 for item in compromisos
        if item["fuente"] == "AVA Fideicomiso" and item["estado"] == "PENDIENTE"
    )
    metrics = [
        ("Gastos personales reales 2026", sum(item["ars"] or 0 for item in personal_2026), f"{len(personal_2026)} movimientos"),
        ("Almacén Sol real 2026", sum(item["ars"] or 0 for item in warehouse_2026), f"{len(warehouse_2026)} movimientos"),
        ("Total real Sol 2026", sum(item["ars"] or 0 for item in actual_2026), f"{len(actual_2026)} movimientos"),
        ("Cashflow BBVA pendiente", sum(item["monto"] for item in cashflow if item["estado"] == "PENDIENTE"), f"{sum(item['estado'] == 'PENDIENTE' for item in cashflow)} cuotas"),
        ("AVA Fideicomiso PAGADO (50% Sol)", ava_pagado_sol, "Ya incluido en Gasto Personal Sol — no duplicar"),
        ("AVA Fideicomiso PENDIENTE", None, f"{ava_pendiente_n} cuotas futuras (monto TBD hasta factura)"),
        ("Percepciones USD 2026", sum(item["ars"] or 0 for item in actual_2026 if item["categoria"] == "PERCEPCIONES USD"), f"{sum(item['categoria'] == 'PERCEPCIONES USD' for item in actual_2026)} movimientos"),
        ("MP No Gasto (retiros/deuda/inversión)", sum(item["monto"] for item in no_gasto if item["moneda"] == "ARS"), f"{len(no_gasto)} movimientos — no es gasto, hoja MP No Gasto"),
        ("Movimientos sin fecha", sum(item["ars"] or 0 for item in actual if item["estado"] == "REVISAR FECHA"), f"{sum(item['estado'] == 'REVISAR FECHA' for item in actual)} movimientos a revisar"),
    ]
    for row, values in enumerate(metrics, 5):
        for column, value in enumerate(values, 1):
            summary.cell(row, column, value).border = THIN
        summary.cell(row, 2).number_format = "#,##0"
    summary.cell(7, 1).font = Font(bold=True)
    summary.cell(7, 2).font = Font(bold=True)

    row = 13
    summary.cell(row, 1, "Resumen mensual 2026").font = Font(size=14, bold=True, color=NAVY)
    row += 1
    write_header(summary, row, ["Mes", "Gasto personal real", "Almacén Sol real", "Total real", "Cashflow BBVA pendiente", "Compromiso total"])
    by_personal = defaultdict(float)
    by_warehouse = defaultdict(float)
    by_pending = defaultdict(float)
    for item in personal_2026:
        by_personal[item["anio_mes"]] += item["ars"] or 0
    for item in warehouse_2026:
        by_warehouse[item["anio_mes"]] += item["ars"] or 0
    for item in cashflow:
        if item["estado"] == "PENDIENTE":
            by_pending[str(item["anio_mes"])] += item["monto"]
    start = row + 1
    for month in range(1, 13):
        current_row = start + month - 1
        year_month_value = f"2026-{month:02d}"
        personal_total = by_personal[year_month_value]
        warehouse_total = by_warehouse[year_month_value]
        pending_total = by_pending[year_month_value]
        values = [
            MONTHS[month], personal_total, warehouse_total,
            personal_total + warehouse_total, pending_total,
            personal_total + warehouse_total + pending_total,
        ]
        for column, value in enumerate(values, 1):
            summary.cell(current_row, column, value).border = THIN
        for column in range(2, 7):
            summary.cell(current_row, column).number_format = "#,##0"
    end = start + 11
    total_row = end + 1
    summary.cell(total_row, 1, "TOTAL 2026").font = Font(bold=True)
    for column in range(2, 7):
        letter = get_column_letter(column)
        summary.cell(total_row, column, f"=SUM({letter}{start}:{letter}{end})")
        summary.cell(total_row, column).number_format = "#,##0"
        summary.cell(total_row, column).font = Font(bold=True)

    row = total_row + 3
    summary.cell(row, 1, "Cashflow pendiente 2027").font = Font(size=14, bold=True, color=NAVY)
    row += 1
    write_header(summary, row, ["Año-Mes", "Monto pendiente BBVA"])
    row += 1
    for year_month_value in sorted(key for key, value in by_pending.items() if key.startswith("2027-") and value):
        summary.cell(row, 1, year_month_value)
        summary.cell(row, 2, by_pending[year_month_value])
        summary.cell(row, 2).number_format = "#,##0"
        row += 1

    category_definitions = {
        "SUPERMERCADO": ("Coto, Carrefour, Día y PedidosYa Market", "ALMACÉN SOL"),
        "COMIDA": ("Restaurantes, delivery, cafeterías y verdulería", "ALMACÉN SOL"),
        "RAMON · COMIDA": ("Alimento de Ramón (Kalbby / DR BARF)", "ALMACÉN SOL"),
        "RAMON · VETERINARIO": ("Valeria Mele / El Perro Verde / Leocan — veterinario Ramón", "ALMACÉN SOL"),
        "RAMON · OTROS": ("Paws Pet Market y otros de Ramón (no comida ni veterinario)", "ALMACÉN SOL"),
        "ENTRETENIMIENTO · SALIDAS COMIDA": ("Hotel Madero — salidas comida (extracto BBVA)", "ALMACÉN SOL"),
        "ENTRETENIMIENTO SOCIAL": ("Restaurant Roldan / salidas sociales", "ALMACÉN SOL"),
        "ENTRETENIMIENTO CENAS": ("Cosi Mi Piace — cenas", "ALMACÉN SOL"),
        "ENTRETENIMIENTO · RECITALES": ("Movistar Arena — recitales", "ALMACÉN SOL"),
        "ENTRETENIMIENTO · TEATRO": ("Plateanet — teatro", "ALMACÉN SOL"),
        "ENTRETENIMIENTO · CINE": ("Hoyts — cine", "ALMACÉN SOL"),
        "COMIDA · BARRITAS PROTEICAS": ("Enerfit — barritas proteicas", "ALMACÉN SOL"),
        "STREAMING TV": ("Directv streaming", "ALMACÉN SOL"),
        "COMBUSTIBLE": ("Shell, YPF y estaciones de servicio", "ALMACÉN SOL"),
        "FARMACIA": ("Farmacity — Almacen Detalle CIUDAD", "ALMACÉN SOL"),
        "REPUESTOS": ("Sodastream y repuestos de casa", "ALMACÉN SOL"),
        "ENTRETENIMIENTO": ("Otros entretenimiento almacén", "ALMACÉN SOL"),
        "PERCEPCIONES USD": ("RG 5617, IVA RG 4240 e IIBB por consumos en USD", "GASTO PERSONAL"),
        "AVA FIDEICOMISO": ("50% de Sol de cuotas AVA", "GASTO PERSONAL"),
        "EVENTOS DEPORTIVOS - RIVER": ("Cuota social CARP / Tu Lugar en el Monumental", "GASTO PERSONAL"),
        "PELUQUERÍA": ("Newplamier y peluquería", "GASTO PERSONAL"),
        "VESTIMENTA": ("Longchamp, Ceferina, Bokenvoreno, Levis, Jackie Smith, Itsclassic, Analia Roay, Jazmin Chebar, Aylen Difeo", "GASTO PERSONAL"),
        "FUNERAL": ("Jardin Del Pilar", "GASTO PERSONAL"),
        "ROPA": ("Indumentaria, calzado, joyería y Tienda River", "GASTO PERSONAL"),
        "SALUD · Vitaminas": ("Promarine / vitaminas y antioxidantes", "GASTO PERSONAL"),
        "SALUD": ("Laboratorios y salud (no farmacia ni Promarine)", "GASTO PERSONAL"),
        "FARMACIA SALUD": ("Farmacia Maure y Farmacia Selma", "GASTO PERSONAL"),
        "TRANSPORTE": ("Cabify, Uber (incl. Ebanx), Didi y Emova", "GASTO PERSONAL"),
        "AUTO": ("ACA y gastos de auto no combustible", "GASTO PERSONAL"),
        "MASCOTAS": ("Veterinaria y accesorios; no comida de Ramón", "GASTO PERSONAL"),
        "MERCADO LIBRE": ("Compras en Mercado Libre", "GASTO PERSONAL"),
        "SUSCRIPCIONES Y TECNOLOGÍA": ("Apple, Cursor, Claude, Gemini, Vercel y similares", "GASTO PERSONAL"),
        "TELECOMUNICACIONES": ("Tuenti y servicios personales", "GASTO PERSONAL"),
        "TRAÍDO DE MIAMI": ("Importadora — Servicios de Pagos Virtuales SA", "GASTO PERSONAL"),
        "USA BOX": ("Importadora — Carenzo / Federico Martinez Escariz", "GASTO PERSONAL"),
        "EDUCACIÓN": ("Universidad y formación", "GASTO PERSONAL"),
        "DONACIONES": ("Donaciones personales", "GASTO PERSONAL"),
        "SEGUROS / RETIRO": ("Zurich y seguros/retiro personales", "GASTO PERSONAL"),
        "COMPRAS ONLINE": ("Amazon y compras online fuera de Mercado Libre", "GASTO PERSONAL"),
        "OTROS PERSONALES": ("Gastos sin categoría más específica", "GASTO PERSONAL"),
    }
    categories_ws.sheet_view.showGridLines = False
    title(categories_ws, "Categorías de gastos Sol", "Totales reales 2026; las cuotas proyectadas se muestran solo en Cashflow")
    write_header(categories_ws, 4, ["Categoría", "Grupo", "Definición", "Movimientos 2026", "Total real 2026 ARS", "% del total"])
    category_count = Counter(item["categoria"] for item in actual_2026)
    category_total = defaultdict(float)
    for item in actual_2026:
        category_total[item["categoria"]] += item["ars"] or 0
    actual_total = sum(category_total.values())
    for row, category in enumerate(sorted(category_definitions), 5):
        definition, group = category_definitions[category]
        values = [
            category, group, definition, category_count[category],
            category_total[category],
            category_total[category] / actual_total if actual_total else 0,
        ]
        for column, value in enumerate(values, 1):
            categories_ws.cell(row, column, value).border = THIN
        categories_ws.cell(row, 5).number_format = "#,##0"
        categories_ws.cell(row, 6).number_format = "0.0%"
    add_table(categories_ws, 4, 4 + len(category_definitions), 6, "TablaCategoriasSol")

    perceptions = [item for item in actual if item["categoria"] == "PERCEPCIONES USD"]
    perception_headers = ["Fecha", "Año-Mes", "Concepto", "Monto ARS", "Medio", "Fuente", "Comentario", "Origen"]
    write_header(perceptions_ws, 1, perception_headers)
    for row, item in enumerate(sorted(perceptions, key=lambda value: (value["fecha"] or date(1900, 1, 1), value["concepto"])), 2):
        values = [
            item["fecha"], item["anio_mes"], item["concepto"], item["ars"],
            item["medio"], item["fuente"], item["comentario"], item["origen"],
        ]
        for column, value in enumerate(values, 1):
            perceptions_ws.cell(row, column, value).border = THIN
        if item["fecha"]:
            perceptions_ws.cell(row, 1).number_format = "YYYY-MM-DD"
        perceptions_ws.cell(row, 4).number_format = "#,##0"
    add_table(perceptions_ws, 1, len(perceptions) + 1, len(perception_headers), "TablaPercepcionesUSD")

    control_ws.sheet_view.showGridLines = False
    title(control_ws, "Control y fuentes", "Trazabilidad y criterios del consolidado")
    write_header(control_ws, 4, ["Control", "Valor", "Observación"])
    controls = [
        ("Archivo canónico fuente", str(SOURCE.relative_to(ROOT)), "Hojas Gasto Personal Sol, Almacen Detalle, Cashflow Cuotas BBVA y Cashflow Compromisos"),
        ("Archivo consolidado", str(OUTPUT.relative_to(ROOT)), "Archivo maestro de gastos de Sol"),
        ("Movimientos personales reales", len(personal), f"ARS {sum(item['ars'] or 0 for item in personal):,.0f}"),
        ("Movimientos Almacén Sol", len(warehouse), f"ARS {sum(item['ars'] or 0 for item in warehouse):,.0f}"),
        ("Cuotas BBVA en cashflow", len(cashflow), f"Pendiente ARS {sum(item['monto'] for item in cashflow if item['estado'] == 'PENDIENTE'):,.0f}"),
        ("Compromisos BBVA+AVA", len(compromisos), f"AVA PAGADO Sol ARS {ava_pagado_sol:,.0f} · AVA PENDIENTE {ava_pendiente_n} cuotas"),
        ("Proyecciones GPS", len(projected), "No se suman como gasto real; son espejo del cashflow BBVA pendiente"),
        ("Movimientos sin fecha", sum(item["estado"] == "REVISAR FECHA" for item in actual), "Completar fecha en próxima revisión"),
        ("Movimientos USD sin ARS", sum(item["ars"] is None and item["usd"] is not None for item in actual), "Se conserva el monto USD"),
    ]
    for row, values in enumerate(controls, 5):
        for column, value in enumerate(values, 1):
            control_ws.cell(row, column, value).border = THIN
    row = 16
    control_ws.cell(row, 1, "Criterios").font = Font(size=14, bold=True, color=NAVY)
    row += 1
    notes = [
        "Movimientos Sol contiene únicamente gastos reales: Gasto Personal Sol + Importe Sol de Almacen Detalle.",
        "Las cuotas PROYECTADO no se duplican en Movimientos Sol; permanecen en Gastos Personales y Cashflow Cuotas BBVA.",
        "Cashflow Compromisos une BBVA pendiente/pagado + AVA Fideicomiso (50% Sol). AVA PAGADO ya está en GPS.",
        "Farmacity → Almacen Detalle · FARMACIA. Sodastream → Almacen Detalle · REPUESTOS. Farmacia Maure/Selma → GPS · FARMACIA SALUD.",
        "PedidosYa Plus → CIUDAD SUSCRIPCIONES (no Gasto Personal Sol ni Almacén).",
        "Los resúmenes oficiales (PDF/CSV) prevalecen sobre texto de chat.",
        "Almacén Sol comprende comida, supermercado, combustible y Ramón (Comida/Veterinario/Otros). OpenAI/ChatGPT → Gasto Personal Sol · SUSCRIPCIONES · IA.",
    ]
    for note in notes:
        control_ws.cell(row, 1, f"• {note}")
        control_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    for sheet in (movements, personal_ws, warehouse_ws, cashflow_ws, compromisos_ws, perceptions_ws):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        widths = {}
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if cell.value is None:
                    continue
                widths[cell.column] = min(max(widths.get(cell.column, 0), len(str(cell.value)) + 2), 55)
                cell.alignment = Alignment(vertical="top")
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = max(11, width)

    movements.column_dimensions["G"].width = 34
    movements.column_dimensions["M"].width = 60
    personal_ws.column_dimensions["K"].width = 60
    warehouse_ws.column_dimensions["J"].width = 55
    cashflow_ws.column_dimensions["K"].width = 55
    compromisos_ws.column_dimensions["C"].width = 40
    compromisos_ws.column_dimensions["K"].width = 60
    categories_ws.column_dimensions["C"].width = 60
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 45
    control_ws.column_dimensions["A"].width = 38
    control_ws.column_dimensions["B"].width = 35
    control_ws.column_dimensions["C"].width = 70

    summary.freeze_panes = "A14"
    summary.page_setup.orientation = "landscape"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 0

    workbook.save(OUTPUT)
    print(f"Created: {OUTPUT}")
    print(f"Actual rows: {len(actual)} | Personal: {len(personal)} | Almacén: {len(warehouse)}")
    print(f"Projected: {len(projected)} | Cashflow BBVA: {len(cashflow)} | Compromisos: {len(compromisos)}")
    print(f"Actual 2026 ARS: {sum(item['ars'] or 0 for item in actual_2026):,.0f}")
    print(f"Pending BBVA ARS: {sum(item['monto'] for item in cashflow if item['estado'] == 'PENDIENTE'):,.0f}")
    print(f"AVA PAGADO Sol ARS: {ava_pagado_sol:,.0f} | AVA PENDIENTE cuotas: {ava_pendiente_n}")
    print(f"MP No Gasto: {len(no_gasto)} movimientos | ARS {sum(item['monto'] for item in no_gasto if item['moneda'] == 'ARS'):,.0f}")


if __name__ == "__main__":
    build()
