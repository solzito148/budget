#!/usr/bin/env python3
"""Cross-check BBVA + MP summaries vs Excel CIUDAD + catalogs."""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from math import ceil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "CIUDAD" / "Gastos_CIUDAD_1132_2026.xlsx"
OUT_JSON = Path(__file__).with_name("audit_crosscheck_2026.json")
OUT_MD = Path(__file__).with_name("audit_crosscheck_2026.md")

EXCLUDE_CATS = {
    "EXCLUIR",
    "DINERO_RETIRADO",
    "TRANSFERENCIA_RECIBIDA",
    "PAGO_TARJETA",
    "IMPUESTOS_AFIP",
    "INVERSION",
    "RESERVA",
    "MP_NO_GASTO",
}

MP_SKIP_DESC = [
    r"DINERO RETIRADO",
    r"RESERVA AUTOM",
    r"DINERO RESERV",
    r"PAGO DE DEUDA",
    r"PAGO DE RESUMEN",
    r"PAGO TARJETA DE CREDITO",
    r"PAGO AUTOMATICO TARJETA",
    r"TRANSFERENCIA RECIBIDA",
    r"INGRESO DE DINERO",
    r"RENDIMIENTOS",
    r"INVERSION BONOS",
    r"INVERSION EMPRESAS",
    r"COMPRA DE DOLAR",
    r"COMPRA DE DOLARES",
    r"VENTA DE",
    r"COMPRA DE D",
    r"PLAZO FIJO",
    r"SANES MARIA SOLE",
    r"MARIA SOLEDAD SANES",
    r"CR[EÉ]DITO.*MERCADO PAGO",
    r"PR[EÉ]STAMO",
    r"DEBITO POR DEUDA",
    r"RECLAMOS Y DEVOLUCIONES",
]


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).date()
        except Exception:
            return None
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def to_int_amount(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(ceil(abs(float(v))))
    s = str(v).strip().replace(" ", "")
    if not s or s.upper().startswith("US$"):
        return None
    s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s
    s = s.replace(",", "")
    try:
        return int(ceil(abs(float(s))))
    except Exception:
        return None


def load_property_workbook(path: Path, label: str, keys):
    if not path.exists():
        return
    wb = load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        for r in range(1, min(sh.max_row or 1, 2000) + 1):
            d = None
            amount = None
            texts = []
            for c in range(1, min(10, sh.max_column or 1) + 1):
                v = sh.cell(r, c).value
                if d is None:
                    td = to_date(v)
                    if td:
                        d = td
                # only numeric cells count as amounts: string columns hold
                # client/partida numbers (e.g. 3016365) that would shadow the real monto
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    a = to_int_amount(v)
                    if a and a >= 100:
                        amount = a if amount is None else max(amount, a)
                if isinstance(v, str) and len(v.strip()) > 2:
                    texts.append(v.strip())
            if d and amount and texts:
                row = {
                    "sheet": f"{label}/{sheet_name}",
                    "fecha": d.isoformat(),
                    "monto": amount,
                    "comercio": " · ".join(texts)[:100],
                    "extra": "",
                }
                keys[(d, amount)].append(row)


def load_excel_keys(wb):
    """Index of (date, amount) -> list of sheet hits; also fuzzy merchant keys."""
    keys = defaultdict(list)  # (date, amount) -> [{sheet, comercio, ...}]

    def add(sheet, d, amount, comercio, extra=None):
        if d is None or amount is None or amount == 0:
            return
        row = {"sheet": sheet, "fecha": d.isoformat(), "monto": amount, "comercio": comercio, "extra": extra or ""}
        keys[(d, amount)].append(row)

    # Gasto Personal Sol
    gps = wb["Gasto Personal Sol"]
    for r in range(22, gps.max_row + 1):
        d = to_date(gps.cell(r, 1).value)
        amount = to_int_amount(gps.cell(r, 4).value)
        comercio = str(gps.cell(r, 3).value or gps.cell(r, 6).value or "")
        add("GPS", d, amount, comercio)

    # Almacen Detalle
    alm = wb["Almacen Detalle"]
    for r in range(4, alm.max_row + 1):
        if str(alm.cell(r, 1).value or "").startswith("TOTAL"):
            continue
        d = to_date(alm.cell(r, 2).value)
        amount = to_int_amount(alm.cell(r, 4).value)
        if amount is None:
            amount = to_int_amount(alm.cell(r, 5).value)
        comercio = str(alm.cell(r, 3).value or "")
        add("Almacen", d, amount, comercio)

    # Detalle CIUDAD (Fecha, Concepto, Descripción, Monto)
    if "Detalle" in wb.sheetnames:
        det = wb["Detalle"]
        for r in range(4, det.max_row + 1):
            d = to_date(det.cell(r, 1).value)
            amount = to_int_amount(det.cell(r, 4).value)
            comercio = f"{det.cell(r, 2).value or ''} {det.cell(r, 3).value or ''}".strip()
            add("Detalle", d, amount, comercio)

    # Other operational sheets that absorb MP/BBVA lines
    for sheet_name in (
        "MP No Gasto",
        "Expensas Consorcios",
        "Dinero Retirado",
        "Inversión",
        "Transferencias recibidas",
        "Pago de Tarjetas",
        "Cashflow Cuotas BBVA",
        "Cashflow Compromisos",
    ):
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        for r in range(1, sh.max_row + 1):
            d = None
            amount = None
            texts = []
            for c in range(1, min(8, sh.max_column or 1) + 1):
                v = sh.cell(r, c).value
                if d is None:
                    d = to_date(v)
                a = to_int_amount(v)
                if a and a >= 100:
                    amount = a if amount is None else max(amount, a)
                if isinstance(v, str) and v.strip():
                    texts.append(v.strip())
            if d and amount and texts:
                add(sheet_name, d, amount, " · ".join(texts)[:100])

    # Other property workbooks
    load_property_workbook(ROOT / "OHIGGINS/Gastos_OHIGGINS_2026.xlsx", "OHIGGINS", keys)
    load_property_workbook(ROOT / "BONORINO/Gastos_BONORINO_2026.xlsx", "BONORINO", keys)
    load_property_workbook(ROOT / "AVA/Gastos_AVA_2026.xlsx", "AVA", keys)
    load_property_workbook(ROOT / "SOL/Gastos_SOL_2026.xlsx", "SOL", keys)

    return keys


def match_excel(keys, d, amount, tol_days=1, tol_amount=1):
    if d is None or amount is None:
        return []
    hits = []
    for dd in range(-tol_days, tol_days + 1):
        for aa in range(-tol_amount, tol_amount + 1):
            try:
                day = d + timedelta(days=dd)
            except Exception:
                continue
            amt = amount + aa
            if amt <= 0:
                continue
            hits.extend(keys.get((day, amt), []))
    return hits


def match_excel_amount_window(keys, d, amount, days=45, tol_amount=1):
    """Match same amount within ±days (utilities dated by emisión vs cobro)."""
    if d is None or amount is None:
        return []
    hits = []
    for (kd, ka), rows in keys.items():
        if abs(ka - amount) <= tol_amount and abs((kd - d).days) <= days:
            hits.extend(rows)
    return hits


def is_utility_or_agg(comercio: str) -> str | None:
    b = norm(comercio)
    if "EDENOR" in b or "LUZ" in b:
        return "utility"
    if "METROGAS" in b or re.search(r"\bGAS\b", b):
        return "utility"
    if "TELECENTRO" in b:
        return "utility"
    if "SUSANA" in b or "LIMPIEZA" in b:
        return "limpieza_agg"
    if "FIDEICOMISO" in b or "AVA PALPA" in b:
        return "ava"
    if "AGUSTINA VISSANI" in b or "AGUS" in b and "PSA" in b:
        return "agus"
    return None


def should_skip_mp(desc: str) -> bool:
    b = norm(desc)
    # Real payments often append "Reserva por gastos…" — only skip pure reserves / no-gasto
    if b.startswith("RESERVA POR GASTOS") or b.startswith("RESERVA AUTOM") or b.startswith("DINERO RESERV"):
        return True
    for pat in MP_SKIP_DESC:
        if pat.startswith(r"RESERVA POR GASTOS"):
            continue
        if re.search(pat, b):
            return True
    return False


def main():
    wb = load_workbook(XLSX, data_only=True)
    excel_keys = load_excel_keys(wb)

    report = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "sources": {},
        "bbva": {},
        "mp_cuenta": {},
        "tc_mp": {},
        "catalog_gaps": {},
        "top_missing": [],
    }

    # ---------- BBVA catalog vs Excel ----------
    bbva_path = ROOT / "CIUDAD/_inbox/tarjetas_bbva/catalogo_bbva_2026.csv"
    bbva_rows = list(csv.DictReader(bbva_path.open(encoding="utf-8-sig")))
    bbva_missing = []
    bbva_matched = 0
    bbva_covered_soft = 0
    bbva_skip = 0
    bbva_2026 = 0

    # index BONIF. CONSUMO rows: Excel often holds the NETO (bruto - bonif)
    bonif_index = defaultdict(list)  # merchant-normalized -> [(date, amount)]
    for r in bbva_rows:
        com = str(r.get("comercio") or "")
        if "BONIF" in com.upper():
            d = to_date(r.get("fecha"))
            a = to_int_amount(r.get("monto_ars"))
            if d and a:
                base = norm(re.sub(r"(?i)BONIF\.?\s*CONSUMO", "", com))[:20]
                bonif_index[base].append((d, a))

    for r in bbva_rows:
        d = to_date(r.get("fecha"))
        if d and d.year != 2026:
            cierre = str(r.get("cierre") or "")
            if "26" not in cierre and (not d or d.year < 2026):
                continue
            if d.year < 2026:
                bbva_skip += 1
                continue
        if d and d.year == 2026:
            bbva_2026 += 1
        amount = to_int_amount(r.get("monto_ars"))
        cat = str(r.get("categoria") or "").upper()
        comercio = str(r.get("comercio") or "")
        if cat in ("EXCLUIR", "PAGO", "CREDITO"):
            bbva_skip += 1
            continue
        if amount is None:
            if r.get("monto_usd"):
                bbva_skip += 1
            continue
        # USD-denominated rows parsed as tiny ARS (e.g. "OPENAI ... USD" $20):
        # the ARS equivalent (USD × TC MEP) is loaded separately in GPS
        if "USD" in comercio and amount <= 250:
            bbva_skip += 1
            continue
        is_bonif = "BONIF" in comercio.upper()
        hits = match_excel(excel_keys, d, amount)
        kind = is_utility_or_agg(comercio + " " + cat)
        if not hits and kind == "utility":
            hits = match_excel_amount_window(excel_keys, d, amount, days=60)
            if hits:
                bbva_covered_soft += 1
        if not hits and not is_bonif:
            # try NETO = bruto - bonif of same merchant within ±10 days
            base = norm(comercio)[:20]
            for bd, ba in bonif_index.get(base, []):
                if abs((bd - d).days) <= 10 and ba < amount:
                    hits = match_excel(excel_keys, d, amount - ba, tol_days=10)
                    if hits:
                        bbva_covered_soft += 1
                        break
        if not hits and is_bonif:
            # bonif row is covered when Excel holds the neto of its bruto sibling
            base = norm(re.sub(r"(?i)BONIF\.?\s*CONSUMO", "", comercio))[:20]
            for r2 in bbva_rows:
                com2 = str(r2.get("comercio") or "")
                if "BONIF" in com2.upper():
                    continue
                if norm(com2)[:20] != base:
                    continue
                d2 = to_date(r2.get("fecha"))
                a2 = to_int_amount(r2.get("monto_ars"))
                if d2 and a2 and abs((d2 - d).days) <= 10 and a2 > amount:
                    if match_excel(excel_keys, d2, a2 - amount, tol_days=10):
                        hits = [{"sheet": "neto"}]
                        bbva_covered_soft += 1
                        break
        if not hits and amount >= 3000:
            # fecha compra (Excel) vs fecha resumen/cobro (catálogo) can differ ~2 weeks
            hits = match_excel_amount_window(excel_keys, d, amount, days=20)
            if hits:
                bbva_covered_soft += 1
        if hits:
            bbva_matched += 1
        else:
            bbva_missing.append(
                {
                    "fecha": d.isoformat() if d else "",
                    "comercio": r.get("comercio"),
                    "monto": amount,
                    "categoria": r.get("categoria"),
                    "destino": r.get("destino"),
                    "card": r.get("card"),
                    "archivo": r.get("archivo"),
                    "hint": kind or "",
                }
            )

    report["bbva"] = {
        "catalog_rows": len(bbva_rows),
        "rows_2026ish": bbva_2026,
        "matched_excel": bbva_matched,
        "matched_soft_utility": bbva_covered_soft,
        "missing_excel": len(bbva_missing),
        "skipped": bbva_skip,
        "missing_sample": sorted(bbva_missing, key=lambda x: -(x["monto"] or 0))[:40],
        "missing_sum": sum(x["monto"] or 0 for x in bbva_missing),
    }

    # ---------- MP cuenta egresos vs catalog + Excel ----------
    mp_mov_path = ROOT / "CIUDAD/_inbox/tarjetas_mp_debito/movimientos_cuenta_mp_2026.csv"
    mp_cat_path = ROOT / "CIUDAD/_inbox/catalogo_mercadopago.csv"
    mp_movs = list(csv.DictReader(mp_mov_path.open(encoding="utf-8-sig")))
    mp_cat = list(csv.DictReader(mp_cat_path.open(encoding="utf-8-sig")))

    # index catalog by (date, amount)
    cat_index = defaultdict(list)
    for r in mp_cat:
        d = to_date(r.get("fecha"))
        a = to_int_amount(r.get("monto"))
        if d and a:
            cat_index[(d, a)].append(r)

    mp_egresos = []
    uncatalogued = []
    catalogued_not_excel = []
    covered_agg = []
    matched_excel_mp = 0
    skipped_mp = 0

    # monthly limpieza totals in Detalle
    limpieza_months = set()
    for (kd, ka), rows in excel_keys.items():
        for row in rows:
            if "LIMPIEZA" in norm(row.get("comercio") or "") and row.get("sheet") == "Detalle":
                limpieza_months.add(kd.month)

    for r in mp_movs:
        if str(r.get("egreso") or "") != "1":
            continue
        d = to_date(r.get("fecha"))
        if not d or d.year != 2026:
            continue
        amount = to_int_amount(r.get("monto_ceil") or r.get("valor"))
        desc = str(r.get("descripcion") or "")
        if amount is None or amount == 0:
            continue
        if should_skip_mp(desc):
            skipped_mp += 1
            continue
        mp_egresos.append(r)

        cat_hits = []
        for dd in range(-1, 2):
            for aa in range(-1, 2):
                day = d + timedelta(days=dd)
                amt = amount + aa
                cat_hits.extend(cat_index.get((day, amt), []))

        if not cat_hits:
            uncatalogued.append(
                {
                    "fecha": d.isoformat(),
                    "monto": amount,
                    "descripcion": desc[:120],
                    "opid": r.get("opid"),
                }
            )
            continue

        cat0 = cat_hits[0]
        cat_name = str(cat0.get("categoria") or "").upper()
        if cat_name in EXCLUDE_CATS or cat_name.startswith("EXCLUIR"):
            skipped_mp += 1
            continue

        hits = match_excel(excel_keys, d, amount)
        kind = is_utility_or_agg(
            f"{cat0.get('comercio')} {cat0.get('categoria')} {cat0.get('comentario')} {desc}"
        )
        if not hits and kind == "utility":
            hits = match_excel_amount_window(excel_keys, d, amount, days=60)
        if not hits and kind == "limpieza_agg" and d.month in limpieza_months:
            covered_agg.append(
                {
                    "fecha": d.isoformat(),
                    "monto": amount,
                    "comercio": cat0.get("comercio"),
                    "note": f"cubierto por Detalle LIMPIEZA mes {d.month}",
                }
            )
            matched_excel_mp += 1
            continue
        if not hits and kind == "ava":
            # AVA full cuota lives in AVA workbook / 50% Sol in GPS — amount won't match 1:1
            covered_agg.append(
                {
                    "fecha": d.isoformat(),
                    "monto": amount,
                    "comercio": cat0.get("comercio"),
                    "note": "AVA cuota (tracker AVA / 50% Sol GPS)",
                }
            )
            matched_excel_mp += 1
            continue
        if not hits and kind == "agus":
            hits = match_excel_amount_window(excel_keys, d, amount, days=30)
        if not hits and ("AGIP" in norm(cat0.get("comercio") or "") or "ABL" in norm(str(cat0.get("categoria") or ""))):
            # varios pagos ABL el mismo día se cargan agregados en Detalle
            same_day = [
                to_int_amount(c.get("monto"))
                for c in mp_cat
                if to_date(c.get("fecha")) == d and "AGIP" in norm(c.get("comercio") or "")
            ]
            total = sum(a for a in same_day if a)
            if len(same_day) > 1 and match_excel(excel_keys, d, total, tol_days=2, tol_amount=2):
                covered_agg.append(
                    {"fecha": d.isoformat(), "monto": amount, "comercio": cat0.get("comercio"),
                     "note": f"cubierto por agregado ABL del día (${total:,})"}
                )
                matched_excel_mp += 1
                continue
        if hits:
            matched_excel_mp += 1
        else:
            catalogued_not_excel.append(
                {
                    "fecha": d.isoformat(),
                    "monto": amount,
                    "descripcion": desc[:100],
                    "categoria": cat0.get("categoria"),
                    "comentario": cat0.get("comentario"),
                    "comercio_cat": cat0.get("comercio"),
                }
            )

    report["mp_cuenta"] = {
        "mov_rows_total": len(mp_movs),
        "egresos_2026_actionable": len(mp_egresos),
        "matched_excel": matched_excel_mp,
        "uncatalogued": len(uncatalogued),
        "catalogued_missing_excel": len(catalogued_not_excel),
        "covered_agg_notes": len(covered_agg),
        "skipped_no_gasto": skipped_mp,
        "uncatalogued_sample": sorted(uncatalogued, key=lambda x: -x["monto"])[:40],
        "uncatalogued_sum": sum(x["monto"] for x in uncatalogued),
        "missing_excel_sample": sorted(catalogued_not_excel, key=lambda x: -x["monto"])[:40],
        "missing_excel_sum": sum(x["monto"] for x in catalogued_not_excel),
        "covered_agg_sample": covered_agg[:15],
    }

    # ---------- TC MP catalog vs Excel ----------
    tc_path = ROOT / "CIUDAD/_inbox/tarjetas_mp/catalogo_tc_mp_2026.csv"
    tc_rows = list(csv.DictReader(tc_path.open(encoding="utf-8-sig")))
    tc_missing = []
    tc_matched = 0
    tc_skip = 0
    for r in tc_rows:
        d = to_date(r.get("fecha_compra") or r.get("fecha_cobro_resumen"))
        amount = to_int_amount(r.get("monto_ars"))
        destino = str(r.get("destino") or "").upper()
        if destino in ("EXCLUIR", "IMPUESTO", "PERCEPCION") or "PERCEPCI" in str(r.get("categoria") or "").upper():
            # still expect percepciones in GPS — don't skip those with category
            pass
        if not amount:
            # USD-only rows (monto_ars 0): the ARS equivalent is loaded via TC MEP in GPS
            tc_skip += 1
            continue
        hits = match_excel(excel_keys, d, amount, tol_days=2)
        # also try cobro date
        d2 = to_date(r.get("fecha_cobro_resumen"))
        if not hits and d2:
            hits = match_excel(excel_keys, d2, amount, tol_days=2)
        if hits:
            tc_matched += 1
        else:
            tc_missing.append(
                {
                    "fecha_compra": (to_date(r.get("fecha_compra")) or date.min).isoformat()
                    if r.get("fecha_compra")
                    else "",
                    "fecha_cobro": (d2.isoformat() if d2 else ""),
                    "descripcion": r.get("descripcion"),
                    "monto": amount,
                    "categoria": r.get("categoria"),
                    "destino": r.get("destino"),
                    "cuota": r.get("cuota"),
                    "periodo": r.get("periodo_resumen"),
                }
            )

    report["tc_mp"] = {
        "catalog_rows": len(tc_rows),
        "matched_excel": tc_matched,
        "missing_excel": len(tc_missing),
        "skipped": tc_skip,
        "missing_sample": sorted(tc_missing, key=lambda x: -(x["monto"] or 0))[:40],
        "missing_sum": sum(x["monto"] or 0 for x in tc_missing),
    }

    # ---------- Catalog quality: REVISAR / empty / PERSONAL vague ----------
    bbva_revisar = [r for r in bbva_rows if str(r.get("categoria") or "").upper() in ("REVISAR", "PENDIENTE", "")]
    mp_revisar = [
        r
        for r in mp_cat
        if str(r.get("categoria") or "").upper() in ("REVISAR", "PENDIENTE", "A_CONFIRMAR", "")
    ]
    report["catalog_gaps"] = {
        "bbva_revisar_or_empty": len(bbva_revisar),
        "bbva_revisar_sample": [
            {"fecha": r.get("fecha"), "comercio": r.get("comercio"), "monto": r.get("monto_ars"), "cat": r.get("categoria")}
            for r in bbva_revisar[:20]
        ],
        "mp_revisar_or_empty": len(mp_revisar),
        "mp_revisar_sample": [
            {"fecha": r.get("fecha"), "comercio": r.get("comercio"), "monto": r.get("monto"), "cat": r.get("categoria")}
            for r in mp_revisar[:20]
        ],
    }

    # Top actionable missing overall
    top = []
    for x in report["bbva"]["missing_sample"][:15]:
        top.append({"fuente": "BBVA", **x, "label": x.get("comercio")})
    for x in report["mp_cuenta"]["uncatalogued_sample"][:15]:
        top.append({"fuente": "MP cuenta (sin catálogo)", **x, "label": x.get("descripcion")})
    for x in report["mp_cuenta"]["missing_excel_sample"][:10]:
        top.append({"fuente": "MP catálogo→Excel", **x, "label": x.get("comercio_cat") or x.get("descripcion")})
    for x in report["tc_mp"]["missing_sample"][:10]:
        top.append({"fuente": "TC MP", **x, "label": x.get("descripcion")})
    top = sorted(top, key=lambda x: -(x.get("monto") or 0))[:50]
    report["top_missing"] = top

    report["sources"] = {
        "excel": str(XLSX),
        "bbva_catalog": str(bbva_path),
        "mp_movimientos": str(mp_mov_path),
        "mp_catalog": str(mp_cat_path),
        "tc_mp_catalog": str(tc_path),
    }

    OUT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    # Markdown summary
    lines = [
        f"# Auditoría cruzada BBVA / MP vs Excel — {report['generated']}",
        "",
        "## Resumen",
        f"- **BBVA catálogo:** {report['bbva']['matched_excel']} en Excel (soft utilidades: {report['bbva'].get('matched_soft_utility', 0)}) · **{report['bbva']['missing_excel']} faltan** (${report['bbva']['missing_sum']:,})",
        f"- **MP cuenta (egresos 2026):** {report['mp_cuenta']['matched_excel']} en Excel · **{report['mp_cuenta']['uncatalogued']} sin catálogo** (${report['mp_cuenta']['uncatalogued_sum']:,}) · **{report['mp_cuenta']['catalogued_missing_excel']} catalogados sin Excel** (${report['mp_cuenta']['missing_excel_sum']:,}) · agregados cubiertos: {report['mp_cuenta'].get('covered_agg_notes', 0)}",
        f"- **TC Mercado Pago:** {report['tc_mp']['matched_excel']} en Excel · **{report['tc_mp']['missing_excel']} faltan** (${report['tc_mp']['missing_sum']:,})",
        f"- Catálogos a revisar (vacío/REVISAR): BBVA {report['catalog_gaps']['bbva_revisar_or_empty']} · MP {report['catalog_gaps']['mp_revisar_or_empty']}",
        "",
        "## Top faltantes por monto",
    ]
    for i, x in enumerate(top[:30], 1):
        lines.append(
            f"{i}. [{x.get('fuente')}] {x.get('fecha') or x.get('fecha_compra') or ''} · ${x.get('monto'):,} · {str(x.get('label') or '')[:80]}"
        )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(OUT_MD.read_text(encoding="utf-8"))
    print(f"\nJSON → {OUT_JSON}")


if __name__ == "__main__":
    main()
